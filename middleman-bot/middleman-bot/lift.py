
from enum import Enum
from typing import List
from telegram import PhotoSize

class LiftState(Enum):
    NONE = -1
    WAREHOUSE = 0
    SHORE = 1
    OPENING = 2
    SITE = 3
    MISSING = 4
    READY = 5


class Lift:
    def __init__(self, id = -1, photo = None, state = LiftState.NONE,
                site = None, opening = None, note = None, users = []):
        self.id = id
        self.photo = photo
        self.state = state
        self.site = site
        self.opening = opening
        self.note = note
        self.users = users

    def clear(self):
        self.id = None
        self.photo = None
        self.state = LiftState.NONE
        self.site = None
        self.opening = None
        self.note = None
        self.users.clear()

import openpyxl

def load_lifts(lift_list : List):
    wb = openpyxl.load_workbook('lifts.xlsx')
    
    lifts_sheet = wb['lifts']


    i = 1

    while True:
        i += 1
        

        id = lifts_sheet.cell(row = i, column = 1).value

        if (id == None):
            break

        photo = 'https://telegram.org/img/t_logo.png'
        state = lifts_sheet.cell(row = i, column = 2).value
        site = lifts_sheet.cell(row = i, column = 3).value
        opening = lifts_sheet.cell(row = i, column = 4).value
        note = lifts_sheet.cell(row = i, column = 5).value

        string_users =  lifts_sheet.cell(row = i, column = 6).value

        users : list = []

        if (string_users != None):
            users = list(string_users.split(":"))
                
        lift_list.append(Lift(id, photo, state, site, opening, note, users))

def add_lift(lift : Lift):
    wb = openpyxl.load_workbook('lifts.xlsx')
    lifts_sheet = wb['lifts']

    r = lift.id + 3

    lifts_sheet.cell(row = r, column = 1).value = lift.id
    lifts_sheet.cell(row = r, column = 2).value = lift.state.name
    lifts_sheet.cell(row = r, column = 3).value = lift.site
    lifts_sheet.cell(row = r, column = 4).value = lift.opening
    lifts_sheet.cell(row = r, column = 5).value = lift.note

    wb.save('lifts.xlsx')

    print("Lift added!")

def modify_lift(lift : Lift):
    wb = openpyxl.load_workbook('lifts.xlsx')
    lifts_sheet = wb['lifts']

    lifts_sheet.cell(row = id, column = 1).value = lift.id
    lifts_sheet.cell(row = id, column = 2).value = lift.state.name
    lifts_sheet.cell(row = id, column = 3).value = lift.site
    lifts_sheet.cell(row = id, column = 4).value = lift.opening
    lifts_sheet.cell(row = id, column = 5).value = lift.note

    wb.save('lifts.xlsx')

    print("Lift modified!")

def modify_lift_state(state : LiftState, id):
    wb = openpyxl.load_workbook('lifts.xlsx')
    lifts_sheet = wb['lifts']

    r = int(id) + 3

    lifts_sheet.cell(row = r, column = 2).value = state.name

    wb.save('lifts.xlsx')

def modify_lift_users(users : List, id):
    wb = openpyxl.load_workbook('lifts.xlsx')
    lifts_sheet = wb['lifts']

    r = int(id) + 3

    print("ID: " + str(r))
    print(users)

    list_str = ':'.join([str(elem) for elem in users])

    lifts_sheet.cell(row = r, column = 6).value = list_str

    wb.save('lifts.xlsx')

def save_lifts(lift_list : List):
    wb = openpyxl.load_workbook('lifts.xlsx')


