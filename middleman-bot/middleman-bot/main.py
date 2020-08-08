
# MORJES TÄSSÄ MUN MUUTOS
#Let's see if this works
# LIT 

""" TODO """

""" _________ || Main || _________ """
""" - Lift preview function [!] """
""" - Ping [!] """

""" _________ || Add || _________ """
""" - Recognize nicnames for sites & openings """
""" - Lift folders for data and photos """
""" - Conversation state shortcuts and notes """
""" - Clear past Callback keyboards """
""" - Multiple photos and scroll buttons (!) """
""" - User choice feedback """
""" - User creation and setup [!] """
""" - Return lift """
""" - Start command for setup [!] """

""" _________ || Update || _________ """
""" - File for helper functions """
""" - Non case sensitive keywords [!] """
""" - Better folder path init """
""" - New lift save system """

""" _________ || Code fix || _________ """
""" - COMMENTS """
""" - Better var names """
""" - Better state names """
""" - Element/Item id to standard """

""" _________ || Optimaze || _________ """
""" - Workbook as argument """

""" _________ || Polish || _________ """
""" - User friendly conversation quide """
""" - Suqqest next state when updating it and highlight current one"""

""" _________ || BUGS || _________ """

""" _________ || Future || _________ """
""" - Software (E) """
""" - Shipments, Lifts, Returns """
""" - Templates (Requests) """
""" - Profiles """



import logging
from telegram import InlineKeyboardButton, InlineKeyboardMarkup, Update
from telegram.ext import Updater, CommandHandler, MessageHandler, Filters, BaseFilter, CallbackContext, CallbackQueryHandler, CallbackContext, ConversationHandler
from pathlib import Path

from lift import Lift, LiftState, load_lifts, add_lift
from follow import Follow
from message_handlers import reply_test, reply_text_default, reply_photo, reply_site, reply_opening, reply_lift, reply_note, reply_user, follow_site, combine
from message_handlers import ConversationState
from callback_query_handlers import preview_button, state_edit_button, follow_site_button, ping_button

from enums import BCD, UD, BD

# Enable logging
logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                    level=logging.INFO)
logger = logging.getLogger(__name__)

import os
print(os.getcwd())

import openpyxl
workbook_file = openpyxl.load_workbook('file.xlsx')
workbook_lifts = openpyxl.load_workbook('lifts.xlsx')


def test_print(update : Update, context : CallbackContext):
    print("Test Print:")
    print("message id: ", update.message.message_id)
    print("chat id: ", update.message.chat.id)


def main():
    
    """Start the bot."""

    updater = Updater("1182411075:AAGsO0gh6609YJTeGa09CBRAZePXm6m5Ivo", use_context=True)

    # Get the dispatcher to register handlers
    dp = updater.dispatcher


    lift_list = []
    load_lifts(lift_list)



    dp.bot_data[BD.LIFT_LIST] = lift_list

    print(workbook_file.sheetnames)

    site_code_sheet = workbook_file['sites']
    opening_code_sheet = workbook_file['openings']
    user_sheet = workbook_file['users']

    follows_sheet = workbook_lifts['follows']


    i = 0
    sites = []
    openings = []

    """Initialize"""

    # Todo: Multiple loops
    while True:
        i += 1
        # Get value from sheet cell
        site_code = site_code_sheet.cell(row = i, column = 1).value
        opening_code = opening_code_sheet.cell(row = i, column = 1).value

        # If all cells empty
        if (site_code == None and opening_code == None):
            break

        # Add value to list
        if (site_code != None):
            sites.append(site_code)
        if (opening_code != None):
            openings.append(opening_code)
    
    i = 1
    users = {}
    while True:
        i += 1

        user_id = user_sheet.cell(row = i, column = 1).value
        user_name = user_sheet.cell(row = i, column = 2).value

        if (user_id == None):
            break

        users[user_name] = user_id

    i = 1
    follow = {}
    while True:
        i += 1

        follow_name = follows_sheet.cell(row = i, column = 1).value

        if (follow_name == None):
            break

        follow_users = []

        j = 1
        while True:
            j += 1

            user = follows_sheet.cell(row = i, column = j).value

            if (user == None):
                break

            follow_users.append(user)

        follow[follow_name] = follow_users


    
    dp.bot_data[BD.USER_DICT] = users
    dp.bot_data[BD.FOLLOW_DICT] = follow

    print(sites)
    print(openings)
    print(users)
    print(follow)

    #TEMP
    last_id = lift_list[-1].id

    dp.bot_data[BD.LAST_ID] = last_id

    follow_site_conversation = ConversationHandler(
        entry_points = [MessageHandler(Filters.text(sites), follow_site)],
        states = {
            ConversationState.FOLLOW_SITE: [CallbackQueryHandler(follow_site_button)]
            },
        fallbacks = [MessageHandler(Filters.text, reply_text_default)])

    new_lift_conversation = ConversationHandler(
        entry_points = [MessageHandler(Filters.photo, reply_photo)],
        states = {
            ConversationState.SITE: [MessageHandler(Filters.text(sites), reply_site)],
            ConversationState.OPENING: [MessageHandler(Filters.text(openings), reply_opening)],
            ConversationState.NOTE: [MessageHandler(Filters.text, reply_note)],
            ConversationState.PREVIEW: [CallbackQueryHandler(preview_button)]
            },
        fallbacks = [MessageHandler(Filters.text, reply_text_default)]
        )

    edit_shipment_conversation = ConversationHandler(
        entry_points = [MessageHandler(Filters.regex(r'ST'), reply_lift)],
        states = {
            ConversationState.EDIT_SHIPMENT: [CallbackQueryHandler(state_edit_button)],
            ConversationState.EDIT_SHIPMENT_LINK: [MessageHandler(Filters.text(users.keys()), reply_user),
                                                   CallbackQueryHandler(state_edit_button)],
            ConversationState.EDIT_SHIPMENT_PING: [CallbackQueryHandler(ping_button)]
            },
        fallbacks = [MessageHandler(Filters.text, reply_text_default)])



    dp.add_handler(new_lift_conversation)
    dp.add_handler(follow_site_conversation)
    dp.add_handler(edit_shipment_conversation)

    dp.add_handler(MessageHandler(Filters.text("testi"), reply_test))
    dp.add_handler(MessageHandler(Filters.text, reply_text_default))


    # Start the Bot
    updater.start_polling()

    updater.idle()

    print("TEST")



if __name__ == '__main__':
    main()
